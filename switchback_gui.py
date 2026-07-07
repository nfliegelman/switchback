#!/usr/bin/env python3
"""
Switchback - CustomTkinter (modern, dark) edition
=================================================================
A modern windowed app for pulling recreation.gov backcountry/wilderness
permit availability. Double-click "Switchback.bat" to launch.

Needs Python 3.8+ with tkinter (in the python.org Windows installer by
default). Auto-installs `customtkinter` and `openpyxl` on first run; if that
ever fails it falls back to a plain CSV / explains what to do.

Exports are written to a `permit_exports` subfolder next to this file.
Data source: recreation.gov.
"""

import csv
import json
import os
import sys
import subprocess
import calendar as calmod
import threading
import queue
import urllib.parse
import urllib.request
import urllib.error
import time
import re as _re
import html as _html
from datetime import date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import tkinter as tk
from tkinter import ttk, messagebox

BASE = "https://www.recreation.gov"
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")
PLACEHOLDER_THRESHOLD = 100_000
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXPORT_DIR = os.path.join(SCRIPT_DIR, "permit_exports")


# ============================ dependency bootstrap ========================
def _ensure(pkg, import_name=None):
    name = import_name or pkg
    try:
        return __import__(name)
    except ImportError:
        pass
    try:
        subprocess.run([sys.executable, "-m", "pip", "install", "--quiet", pkg],
                       check=True, timeout=240,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return __import__(name)
    except Exception:
        return None


# CustomTkinter must exist before the classes below are defined (they subclass
# its widgets), so resolve it now - first run may pause briefly to install it.
ctk = _ensure("customtkinter")
if ctk is None:
    _r = tk.Tk(); _r.withdraw()
    messagebox.showerror(
        "Setup needed",
        "Couldn't install the 'customtkinter' component.\n\n"
        "This usually means no internet on first run, or pip is unavailable.\n"
        "Connect to the internet and run again, or install it manually:\n\n"
        "    python -m pip install customtkinter")
    raise SystemExit(1)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")


# ============================ network layer ===============================
def _get_json(url, retries=3, timeout=30, backoff=0.6):
    req = urllib.request.Request(url, headers={"User-Agent": UA,
                                               "Accept": "application/json"})
    last = None
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return json.load(r), None
        except urllib.error.HTTPError as e:
            return None, f"HTTP {e.code}"
        except Exception as e:
            last = str(e)
            time.sleep(backoff * (attempt + 1))
    return None, last


def search_permits(query, size=15):
    url = (f"{BASE}/api/search?q={urllib.parse.quote(query)}"
           f"&entity_type=permit&size={size}")
    data, err = _get_json(url)
    if err:
        return [], err
    out = []
    for r in data.get("results", []):
        if r.get("entity_type") != "permit":
            continue
        out.append({"permit_id": str(r.get("entity_id")),
                    "name": r.get("name") or "",
                    "park": r.get("parent_name", "") or ""})
    out.sort(key=lambda r: (0 if "permit" in r["name"].lower() else 1, r["name"]))
    return out, None


# --- description parsing: trail + lake info out of the HTML blurb ----------
_ON_TRAIL_RE = _re.compile(r"on the ([^.]+)\.", _re.I)
_ON_TRAIL_FALLBACK = _re.compile(r"on the ([^.]+?Trail[s]?)\b", _re.I)
_LAKE_RE = _re.compile(r"\b(lakes?|tarns?)\b", _re.I)


def _strip_html(s):
    s = _re.sub(r"<[^>]+>", " ", s or "")
    return _html.unescape(_re.sub(r"\s+", " ", s)).strip()


def parse_description(name, desc):
    text = _strip_html(desc)
    if not text:
        return "", ("Y" if _LAKE_RE.search(name or "") else "")
    m = _ON_TRAIL_RE.search(text) or _ON_TRAIL_FALLBACK.search(text)
    on_trail = m.group(1).strip().replace(" and ", "; ") if m else ""
    feature = _re.split(r"closest\s+trailhead", text, flags=_re.I)[0]
    water = "Y" if _LAKE_RE.search((name or "") + " " + feature) else ""
    return on_trail, water


def get_divisions(permit_id):
    data, err = _get_json(f"{BASE}/api/permitcontent/{permit_id}")
    if err:
        return None, err
    payload = data.get("payload", {}) or {}
    divs = payload.get("divisions") or {}
    out = []
    for v in divs.values():
        on_trail, water = parse_description(v.get("name"), v.get("description"))
        out.append({"id": v.get("id"), "name": v.get("name"),
                    "type": v.get("type"),
                    "is_group": "(Group Site)" in (v.get("name") or ""),
                    "district": v.get("district") or "",
                    "on_trail": on_trail, "water_lake_flag": water})
    return {"permit_name": payload.get("name", ""), "divisions": out}, None


def fetch_division_month(permit_id, division_id, year, month):
    url = (f"{BASE}/api/permititinerary/{permit_id}/division/{division_id}"
           f"/availability/month?month={int(month)}&year={int(year)}"
           f"&commercial=false")
    data, err = _get_json(url)
    if err:
        return None, err
    payload = data.get("payload", {}) or {}
    per_date = {}
    for _qt, datemap in (payload.get("quota_type_maps") or {}).items():
        for d, cell in datemap.items():
            cur = per_date.setdefault(d, {"remaining": 0, "total": 0,
                                          "hidden": True, "walkup": False})
            cur["remaining"] += cell.get("remaining", 0)
            cur["total"] += cell.get("total", 0)
            cur["hidden"] = cur["hidden"] and cell.get("is_hidden", False)
            cur["walkup"] = cur["walkup"] or bool(cell.get("show_walkup"))
    if not per_date:
        for d, avail in (payload.get("bools") or {}).items():
            per_date[d] = {"remaining": 1 if avail else 0, "total": None,
                           "hidden": not avail, "walkup": False}
    return per_date, None


def classify_status(remaining, total, walkup, hidden):
    """Primary disposition of a camp-night. walkup is tracked separately too."""
    if hidden:
        return "Hidden"
    if remaining and remaining > 0:
        return "Reservable"
    if (total == 0 or total is None) and walkup:
        return "Walk-up only"
    if total and total > 0 and (remaining or 0) == 0:
        return "Full"
    if (total == 0 or total is None) and not walkup:
        return "Not released"
    return ""


def daterange(s, e):
    d = s
    while d <= e:
        yield d
        d += timedelta(days=1)


# ====================== output writers (xlsx / csv) =======================
def write_csv(rows, headers, path):
    with open(path, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(headers)
        w.writerows(rows)


def write_xlsx(rows, headers, path, idx):
    """Styled workbook. idx maps logical columns to 1-based positions:
    {'status','remaining','total','pct','walkup','lake'}."""
    opx = _ensure("openpyxl")
    if opx is None:
        alt = os.path.splitext(path)[0] + ".csv"
        write_csv(rows, headers, alt)
        return alt, False
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Availability"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = PatternFill("solid", fgColor="2F6F4E")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(list(r))
    n = len(rows)
    for c, h in enumerate(headers, 1):
        longest = max([len(str(h))] + [len(str(row[c - 1])) for row in rows]) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 8), 42)
    if n:
        L = get_column_letter
        if idx.get("pct"):
            col = L(idx["pct"])
            for r in range(2, n + 2):
                ws.cell(r, idx["pct"]).number_format = "0%"
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                               mid_type="num", mid_value=0.5, mid_color="FFEB84",
                               end_type="num", end_value=1, end_color="63BE7B"))
        if idx.get("lake"):
            col = L(idx["lake"])
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                CellIsRule(operator="equal", formula=['"Y"'],
                           fill=PatternFill("solid", fgColor="BDE0FE")))
        if idx.get("walkup"):
            col = L(idx["walkup"])
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                CellIsRule(operator="equal", formula=['"Y"'],
                           fill=PatternFill("solid", fgColor="E3D7F4")))
        if idx.get("remaining"):
            col = L(idx["remaining"])
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                CellIsRule(operator="lessThanOrEqual", formula=["2"],
                           fill=PatternFill("solid", fgColor="FFE0B3")))
        if idx.get("status"):
            col = L(idx["status"])
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                CellIsRule(operator="equal", formula=['"Walk-up only"'],
                           fill=PatternFill("solid", fgColor="E3D7F4")))
            ws.conditional_formatting.add(
                f"{col}2:{col}{n + 1}",
                CellIsRule(operator="equal", formula=['"Full"'],
                           fill=PatternFill("solid", fgColor="ECECEC")))
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    if n >= 1:
        # Make it a real Excel Table → filter buttons on every column + banded
        # rows, and lets the user add real Slicers in one click.
        from openpyxl.worksheet.table import Table, TableStyleInfo
        try:
            table = Table(displayName="AvailabilityTable",
                          ref=f"A1:{last_col}{n + 1}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True,
                showColumnStripes=False)
            ws.add_table(table)
        except Exception:
            ws.auto_filter.ref = f"A1:{last_col}{n + 1}"
    else:
        ws.auto_filter.ref = f"A1:{last_col}1"
    wb.save(path)
    return path, True


# ============================ calendar popup ==============================
class CalendarPopup(ctk.CTkToplevel):
    """Clickable month calendar, Sunday-first. Sets target_var to YYYY-MM-DD."""

    def __init__(self, parent, target_var, initial=None):
        super().__init__(parent)
        self.title("Pick a date")
        self.resizable(False, False)
        self.target_var = target_var
        self.transient(parent)
        self.after(120, self.grab_set)

        today = date.today()
        if initial:
            try:
                today = date.fromisoformat(initial)
            except ValueError:
                pass
        self.year, self.month = today.year, today.month

        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(padx=12, pady=(12, 4), fill="x")
        ctk.CTkButton(hdr, text="‹", width=34, command=self.prev_month).pack(side="left")
        self.lbl = ctk.CTkLabel(hdr, text="", width=170,
                                font=ctk.CTkFont(size=14, weight="bold"))
        self.lbl.pack(side="left", padx=6)
        ctk.CTkButton(hdr, text="›", width=34, command=self.next_month).pack(side="left")

        self.grid_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.grid_frame.pack(padx=12, pady=(0, 12))
        self.draw()

    def draw(self):
        for w in self.grid_frame.winfo_children():
            w.destroy()
        self.lbl.configure(text=f"{calmod.month_name[self.month]} {self.year}")
        for i, d in enumerate(["Su", "Mo", "Tu", "We", "Th", "Fr", "Sa"]):
            ctk.CTkLabel(self.grid_frame, text=d, width=36,
                         text_color="#8E9197").grid(row=0, column=i, pady=(0, 2))
        cal = calmod.Calendar(firstweekday=6).monthdayscalendar(self.year, self.month)
        today = date.today()
        for r, week in enumerate(cal, start=1):
            for c, day in enumerate(week):
                if day == 0:
                    continue
                dt = date(self.year, self.month, day)
                is_today = dt == today
                b = ctk.CTkButton(
                    self.grid_frame, text=str(day), width=36, height=30,
                    corner_radius=8,
                    fg_color=("#1F6AA5" if is_today else "transparent"),
                    hover_color="#144870",
                    command=lambda d=dt: self.choose(d))
                b.grid(row=r, column=c, padx=1, pady=1)

    def prev_month(self):
        self.month -= 1
        if self.month == 0:
            self.month, self.year = 12, self.year - 1
        self.draw()

    def next_month(self):
        self.month += 1
        if self.month == 13:
            self.month, self.year = 1, self.year + 1
        self.draw()

    def choose(self, dt):
        self.target_var.set(dt.isoformat())
        self.destroy()


# ================================ main app ================================
PCT_DISPLAY_INDEX = 7  # 0-based position of pct in a row tuple (see headers)


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Switchback")
        self.geometry("940x720")
        self.minsize(860, 640)

        self.permit_id = None
        self.permit_name = None
        self.divisions = []
        self.search_results = []
        self.msg_q = queue.Queue()
        self.last_path = None

        self._build()

    def _build(self):
        wrap = ctk.CTkFrame(self, fg_color="transparent")
        wrap.pack(fill="both", expand=True, padx=14, pady=12)

        # Step 1: park / permit
        f1 = ctk.CTkFrame(wrap, corner_radius=12)
        f1.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(f1, text="1.  Park / permit",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=14, pady=(12, 6))
        r1 = ctk.CTkFrame(f1, fg_color="transparent"); r1.pack(fill="x", padx=14)
        self.search_var = tk.StringVar()
        e = ctk.CTkEntry(r1, textvariable=self.search_var,
                         placeholder_text="search a park, e.g. rainier wilderness")
        e.pack(side="left", fill="x", expand=True)
        e.bind("<Return>", lambda _e: self.do_search())
        ctk.CTkButton(r1, text="Search", width=90, command=self.do_search).pack(side="left", padx=(8, 0))

        r2 = ctk.CTkFrame(f1, fg_color="transparent"); r2.pack(fill="x", padx=14, pady=(8, 0))
        self.permit_combo = ctk.CTkComboBox(r2, values=["(search results appear here)"],
                                            command=self.on_permit_selected)
        self.permit_combo.pack(side="left", fill="x", expand=True)

        r3 = ctk.CTkFrame(f1, fg_color="transparent"); r3.pack(fill="x", padx=14, pady=(8, 0))
        ctk.CTkLabel(r3, text="or paste a permit ID:").pack(side="left")
        self.id_var = tk.StringVar()
        ctk.CTkEntry(r3, textvariable=self.id_var, width=120).pack(side="left", padx=8)
        ctk.CTkButton(r3, text="Load", width=70, command=self.load_by_id).pack(side="left")

        self.permit_status = ctk.CTkLabel(f1, text="No permit loaded.",
                                          text_color="#E06A6A")
        self.permit_status.pack(anchor="w", padx=14, pady=(8, 12))

        # Step 2: dates
        f2 = ctk.CTkFrame(wrap, corner_radius=12)
        f2.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(f2, text="2.  Dates",
                     font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=14, pady=(12, 6))
        rd = ctk.CTkFrame(f2, fg_color="transparent"); rd.pack(fill="x", padx=14, pady=(0, 12))
        self.start_var = tk.StringVar()
        self.end_var = tk.StringVar()
        ctk.CTkLabel(rd, text="First night:").pack(side="left")
        ctk.CTkEntry(rd, textvariable=self.start_var, width=130).pack(side="left", padx=(6, 4))
        ctk.CTkButton(rd, text="Pick", width=60,
                      command=lambda: self.pick(self.start_var)).pack(side="left")
        ctk.CTkLabel(rd, text="   Last night:").pack(side="left", padx=(12, 0))
        ctk.CTkEntry(rd, textvariable=self.end_var, width=130).pack(side="left", padx=(6, 4))
        ctk.CTkButton(rd, text="Pick", width=60,
                      command=lambda: self.pick(self.end_var)).pack(side="left")
        ctk.CTkLabel(rd, text="(leave last night empty for one night)",
                     text_color="#8E9197").pack(side="left", padx=(12, 0))

        # Step 3: options + run
        f3 = ctk.CTkFrame(wrap, corner_radius=12)
        f3.pack(fill="x", pady=(0, 10))
        r3o = ctk.CTkFrame(f3, fg_color="transparent"); r3o.pack(fill="x", padx=14, pady=12)
        self.only_avail = tk.BooleanVar(value=True)
        self.skip_group = tk.BooleanVar(value=True)
        ctk.CTkCheckBox(r3o, text="Only show available (incl. walk-up)",
                        variable=self.only_avail).pack(side="left")
        ctk.CTkCheckBox(r3o, text="Skip group sites",
                        variable=self.skip_group).pack(side="left", padx=18)
        self.run_btn = ctk.CTkButton(r3o, text="Find availability", command=self.run)
        self.run_btn.pack(side="right")

        # progress
        self.prog = ctk.CTkProgressBar(wrap)
        self.prog.set(0)
        self.prog.pack(fill="x", pady=(0, 2))
        self.prog_lbl = ctk.CTkLabel(wrap, text="", text_color="#8E9197")
        self.prog_lbl.pack(anchor="w")

        # results table (ttk.Treeview styled dark to match)
        tbl = ctk.CTkFrame(wrap, corner_radius=12)
        tbl.pack(fill="both", expand=True, pady=(8, 8))
        self._style_tree()
        cols = ("camp", "type", "district", "date", "status", "remaining",
                "total", "pct", "walkup", "lake", "on_trail")
        heads = ("Camp/Zone", "Type", "District", "Date", "Status", "Rem",
                 "Total", "% Left", "Walk-up", "Lake", "On Trail")
        widths = (170, 120, 140, 80, 100, 55, 55, 60, 70, 50, 170)
        self.tree = ttk.Treeview(tbl, columns=cols, show="headings", height=10,
                                 style="Dark.Treeview")
        for c, h, w in zip(cols, heads, widths):
            self.tree.heading(c, text=h)
            anchor = "w" if c in ("camp", "type", "district", "status", "on_trail") else "center"
            self.tree.column(c, width=w, anchor=anchor)
        vsb = ttk.Scrollbar(tbl, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=(8, 0), pady=8)
        vsb.grid(row=0, column=1, sticky="ns", pady=8)
        hsb.grid(row=1, column=0, sticky="ew", padx=8)
        tbl.rowconfigure(0, weight=1); tbl.columnconfigure(0, weight=1)

        bottom = ctk.CTkFrame(wrap, fg_color="transparent"); bottom.pack(fill="x")
        self.save_lbl = ctk.CTkLabel(bottom, text="", text_color="#5FC58A")
        self.save_lbl.pack(side="left")
        self.open_btn = ctk.CTkButton(bottom, text="Open exports folder", width=150,
                                      command=self.open_folder, state="disabled")
        self.open_btn.pack(side="right")

    def _style_tree(self):
        style = ttk.Style()
        try:
            style.theme_use("default")
        except tk.TclError:
            pass
        style.configure("Dark.Treeview", background="#2B2B2B",
                        fieldbackground="#2B2B2B", foreground="#DCE4EE",
                        rowheight=26, borderwidth=0)
        style.configure("Dark.Treeview.Heading", background="#323232",
                        foreground="#A6ABB2", relief="flat",
                        borderwidth=0)
        style.map("Dark.Treeview", background=[("selected", "#1F6AA5")])
        style.map("Dark.Treeview.Heading", background=[("active", "#3A3A3A")])

    # ---- actions ----------------------------------------------------------
    def pick(self, var):
        CalendarPopup(self, var, initial=var.get() or None)

    def do_search(self):
        q = self.search_var.get().strip()
        if not q:
            return
        self.permit_status.configure(text="Searching…", text_color="#8E9197")
        self.update_idletasks()
        results, err = search_permits(q)
        if err:
            messagebox.showerror("Search failed", err); return
        if not results:
            self.permit_status.configure(text="No permits found. Try other words.",
                                         text_color="#E06A6A")
            self.permit_combo.configure(values=["(no results)"]); return
        self.search_results = results
        labels = [f"{r['name']}  -  {r['park']}" if r["park"] else r["name"]
                  for r in results]
        self.permit_combo.configure(values=labels)
        self.permit_combo.set(labels[0])
        if len(results) == 1:
            self.load_permit(results[0]["permit_id"])  # only one match → load it
        else:
            self.permit_status.configure(text="Pick a result from the dropdown.",
                                         text_color="#8E9197")

    def on_permit_selected(self, choice):
        for i, r in enumerate(self.search_results):
            label = f"{r['name']}  -  {r['park']}" if r["park"] else r["name"]
            if label == choice:
                self.load_permit(r["permit_id"]); return

    def load_by_id(self):
        pid = self.id_var.get().strip()
        if not pid.isdigit():
            messagebox.showwarning("Permit ID", "Enter the numeric permit ID, e.g. 4675317.")
            return
        self.load_permit(pid)

    def load_permit(self, pid):
        self.permit_status.configure(text="Loading permit…", text_color="#8E9197")
        self.update_idletasks()
        info, err = get_divisions(pid)
        if err:
            self.permit_status.configure(
                text=f"Couldn't load permit {pid} ({err}). "
                     f"That entry may not be a real bookable permit.",
                text_color="#E06A6A")
            return
        self.permit_id = pid
        self.permit_name = info["permit_name"]
        self.divisions = info["divisions"]
        self.permit_status.configure(
            text=f"✓ Loaded: {self.permit_name}  ({len(self.divisions)} camps/zones)",
            text_color="#5FC58A")

    def run(self):
        if not self.permit_id:
            messagebox.showwarning("Pick a permit", "Load a park/permit first."); return
        try:
            start = date.fromisoformat(self.start_var.get())
        except ValueError:
            messagebox.showwarning("Dates", "Pick a first night."); return
        end_s = self.end_var.get()
        end = start
        if end_s:
            try:
                end = date.fromisoformat(end_s)
            except ValueError:
                end = start
        if end < start:
            start, end = end, start

        divisions = list(self.divisions)
        if self.skip_group.get():
            divisions = [d for d in divisions if not d["is_group"]]

        want_dates = [d.isoformat() for d in daterange(start, end)]
        yms = sorted({(d.year, d.month) for d in daterange(start, end)})
        jobs = [(d, y, m) for d in divisions for (y, m) in yms]

        for r in self.tree.get_children():
            self.tree.delete(r)
        self.save_lbl.configure(text="")
        self.open_btn.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        self.prog.set(0)
        self.prog_lbl.configure(text=f"Fetching {len(jobs)} records…")

        args = (jobs, divisions, want_dates, start, end, self.only_avail.get())
        threading.Thread(target=self._worker, args=args, daemon=True).start()
        self.after(100, self._drain)

    def _worker(self, jobs, divisions, want_dates, start, end, only_avail):
        results, errors, done = {}, 0, 0

        def work(job):
            d, y, m = job
            per, e = fetch_division_month(self.permit_id, d["id"], y, m)
            return d, y, m, per, e

        with ThreadPoolExecutor(max_workers=6) as ex:
            futs = [ex.submit(work, j) for j in jobs]
            for f in as_completed(futs):
                d, y, m, per, e = f.result()
                done += 1
                if e:
                    errors += 1
                else:
                    results[(d["id"], y, m)] = per
                self.msg_q.put(("progress", done / max(len(jobs), 1)))

        rows = []
        for d in divisions:
            for ds in want_dates:
                y, m = int(ds[:4]), int(ds[5:7])
                cell = results.get((d["id"], y, m), {}).get(ds)
                if cell is None:
                    remaining = total = None; hidden = True; walkup = False
                else:
                    remaining, total = cell.get("remaining"), cell.get("total")
                    hidden = cell.get("hidden", False)
                    walkup = cell.get("walkup", False)
                if remaining is not None and remaining >= PLACEHOLDER_THRESHOLD:
                    continue
                status = classify_status(remaining, total, walkup, hidden)
                if status == "Hidden":
                    continue
                if only_avail and status not in ("Reservable", "Walk-up only"):
                    continue
                pct = (remaining / total) if (remaining is not None and total) else None
                rows.append((d["name"], d["type"], d.get("district", ""), ds,
                             status, remaining, total, pct,
                             "Y" if walkup else "", d.get("water_lake_flag", ""),
                             d.get("on_trail", "")))
        rows.sort(key=lambda r: (r[1] or "", r[0] or "", r[3]))

        headers = ["camp_or_zone", "type", "district", "date", "status",
                   "remaining", "total", "percent_remaining", "walkup",
                   "water_lake_flag", "on_trail"]
        idx = {"status": 5, "remaining": 6, "total": 7, "pct": 8,
               "walkup": 9, "lake": 10}

        try:
            os.makedirs(EXPORT_DIR, exist_ok=True)
        except Exception as ex:
            self.msg_q.put(("error", f"Couldn't create exports folder: {ex}")); return
        safe = "".join(c if c.isalnum() else "_"
                       for c in (self.permit_name or self.permit_id))[:40]
        out = os.path.join(EXPORT_DIR, f"availability_{safe}_{start}_to_{end}.xlsx")
        self.msg_q.put(("status", "Saving spreadsheet… (first run may install Excel support)"))
        try:
            saved, styled = write_xlsx(rows, headers, out, idx)
        except Exception as ex:
            self.msg_q.put(("error", str(ex))); return
        self.msg_q.put(("done", (rows, saved, errors, len(jobs), styled)))

    def _drain(self):
        try:
            while True:
                kind, payload = self.msg_q.get_nowait()
                if kind == "progress":
                    self.prog.set(payload)
                    self.prog_lbl.configure(text=f"Fetching… {round(payload * 100)}%")
                elif kind == "status":
                    self.prog_lbl.configure(text=payload)
                elif kind == "error":
                    self.run_btn.configure(state="normal")
                    messagebox.showerror("Error", payload); return
                elif kind == "done":
                    rows, out, errors, njobs, styled = payload
                    self._show_results(rows, out, errors, njobs, styled); return
        except queue.Empty:
            pass
        self.after(100, self._drain)

    def _show_results(self, rows, out, errors, njobs, styled):
        self.run_btn.configure(state="normal")
        self.prog.set(1)
        for r in rows:
            disp = []
            for i, v in enumerate(r):
                if v is None:
                    disp.append("")
                elif i == PCT_DISPLAY_INDEX:
                    disp.append(f"{round(v * 100)}%")
                else:
                    disp.append(v)
            self.tree.insert("", "end", values=disp)
        self.prog_lbl.configure(text=f"Done. {len(rows)} rows."
                                + (f"  ({errors}/{njobs} calls errored)" if errors else ""))
        self.last_path = out
        kind = "Excel file" if styled else "CSV (Excel formatting unavailable)"
        self.save_lbl.configure(text=f"Saved {kind}: permit_exports/{os.path.basename(out)}")
        self.open_btn.configure(state="normal")
        self._open_path(out)  # auto-open the spreadsheet
        if errors and errors == njobs:
            messagebox.showwarning(
                "No data",
                "Every request failed. This permit may use a different "
                "availability model, or the site is temporarily blocking requests.")

    @staticmethod
    def _open_path(path):
        """Open a file or folder in the OS default handler. Never fatal."""
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # noqa
            elif sys.platform == "darwin":
                os.system(f'open "{path}"')
            else:
                os.system(f'xdg-open "{path}"')
        except Exception:
            pass

    def open_folder(self):
        target = EXPORT_DIR if os.path.isdir(EXPORT_DIR) else SCRIPT_DIR
        self._open_path(target)


def main():
    try:
        App().mainloop()
    except Exception as e:
        try:
            r = tk.Tk(); r.withdraw()
            messagebox.showerror("Permit Finder crashed", str(e))
        except Exception:
            print("Error:", e)


if __name__ == "__main__":
    main()
